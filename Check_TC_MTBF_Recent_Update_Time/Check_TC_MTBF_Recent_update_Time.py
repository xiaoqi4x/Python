import csv
import os
from founction import InformationCheck
import requests
import openpyxl
import urllib3
import time


# class Outputer:
ic = InformationCheck()
var = 1
while var == 1:
    # os.remove('C:\zxq\Check_TC_MTBF_Recent_Update_Time\Build.xlsx')
    section_name = ic.section_name()
    # def result_outputer(section_name):
    path = "C:\zxq\Check_TC_MTBF_Recent_Update_Time\Racks.xlsx"
    workbook = openpyxl.load_workbook(path)
    for i in range(len(workbook.sheetnames)):
        sheet_name = workbook.sheetnames[i]
        if section_name == sheet_name:
            sheet_row = workbook[workbook.sheetnames[i]].max_row
            print(sheet_name + ' ' + str(sheet_row))
            for j in range(1, sheet_row + 1):
                sheet_col_1 = workbook[workbook.sheetnames[i]].cell(j, 1).value
                print(sheet_col_1)
                #old:https:https://harts.intel.com/harts6/ListTestsets.do?slaveName=BEJSRTL625.bj.intel.com
                #new:https://oc6web-harts.intel.com/integfe/node/details/BEJSRTL688.bj.intel.com
                sheet_col_2 = workbook[workbook.sheetnames[i]].cell(j, 2).value
                if not sheet_col_2:
                    continue
                rack_rat = str(sheet_col_1)[-23:-13] + '  ' + sheet_col_2 + '  '
                # if requests.get('http://' + sheet_col_1 + ':8080/harts/').status_code == 200:
                urllib3.disable_warnings()
                if requests.get(sheet_col_1, verify=False).status_code == 200:
                    page_content = ic.webpage(str(sheet_col_1))
                    abort_count = page_content.find_all(type='button', text='abort')
                    if len(abort_count) >= 2:
                        print(rack_rat + 'has more than 1 request is pending!!!')
                        continue
                    elif len(abort_count) == 0:
                        print(rack_rat + 'stopped!!!')
                        continue
                    else:
                        test_id = str(abort_count[0])
                        link = 'http://' + str(sheet_col_1)[-23:-13] + ':8080/harts/logs/test_sets/' + test_id[
                                                                                                       123:131] + '/test_engine/'
                        url = link + ic.check_mtbf_release_test_folder(link)
                        if ic.output_flash_txt_link(url) == 'false' or str(
                                ic.find_prepare_folder(url)) == 'false' or ic.check_tc_mtbf_7660_txt(url) == 'false':
                            print(rack_rat + 'cannot ping successfully!!!-1')
                            continue
                        elif str(ic.output_flash_txt_link(url)) == 'still not flashing...':
                            print(rack_rat + 'still not flashing...')
                            continue
                        elif str(ic.find_prepare_folder(url)) == 'still no prepare folder...':
                            print(rack_rat + 'still no prepare folder...')
                            continue
                        else:
                            pre_result = ic.check_rat_combo_from_prepare(url)
                            if not pre_result:
                                pre_result = 'None'
                            print(rack_rat + pre_result + '  ' + ic.check_tc_mtbf_7660_txt(
                                url) + '  ' + ic.output_flash_txt_link(
                                url))
                else:
                    print(rack_rat + 'cannot ping successfully!!!-2')
        else:
            continue








